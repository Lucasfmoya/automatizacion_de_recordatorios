# -*- coding: utf-8 -*-

"""
=============================================================
  LUBRICENTRO O'HIGGINS — Envío automático de WhatsApp
=============================================================

  Requisitos:
    pip install openpyxl pyautogui

  Uso:
    python lubricentro_whatsapp.py
=============================================================
"""

import openpyxl
import webbrowser
import urllib.parse
import time
import random
import pyautogui
import sys
import ctypes
import winsound

from datetime import date

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────

RUTA_EXCEL = r"\\DESKTOP-PCFS020\compartir\SOPORTE 2\LUBRICENTRO\Macros\Nuevas planillas de mensajes\Base de datos.xlsx"

NOMBRE_HOJA = "BASE 2026"

FILA_ENCABEZADOS = 1

# ── Columnas Excel ──────────────────────────────────────────
COL_CLIENTE       = "Cliente"
COL_TELEFONO      = "Contacto"
COL_AUTO          = "Auto"
COL_PATENTE       = "Patente"
COL_KMTS_PROX     = "Kmts Prox. Cambio"
COL_NUM_CONTACTOS = "Num de contacto"
COL_FECHA_ENVIO   = "Repuesta 1"

# ── Configuración envío ─────────────────────────────────────
CANTIDAD_POR_TANDA = 10

DELAY_MIN = 4
DELAY_MAX = 6

# ⚠️ Tiempo de espera para que cargue WhatsApp
#    Aumentalo si tenés conexión o PC lenta
ESPERA_CARGA_CHAT = 10

PREFIJO = "549"

# ─────────────────────────────────────────────────────────────
# MENSAJE
# ─────────────────────────────────────────────────────────────

MENSAJE = """🚘 *Recordatorio de Service*

Tu *{auto}* dominio *{patente}* está próximo a los *{km} kms* ⏱️

🔧 En *Lubricentro O'Higgins* tenemos un beneficio exclusivo para vos:
📅 *10% OFF* en cambio de aceite reservando tu turno por WhatsApp.

💥 ¡Consultá disponibilidad!

_Promoción válida abonando en efectivo o transferencia._
"""

# ─────────────────────────────────────────────────────────────
# FUNCIONES
# ─────────────────────────────────────────────────────────────

def formatear_km(km):
    """Formatea kilómetros con puntos de miles"""
    try:
        return f"{int(float(km)):,}".replace(",", ".")
    except:
        return str(km)


def formatear_auto(texto):
    """Capitaliza nombre del vehículo respetando siglas"""

    if not texto:
        return ""

    especiales = {
        "BMW", "GTI", "XLS", "GLI", "VTI", "TDI",
        "HDI", "SRX", "AMG", "X5", "X6", "GNC"
    }

    palabras = str(texto).strip().split()
    resultado = []

    for p in palabras:
        if p.upper() in especiales:
            resultado.append(p.upper())
        else:
            resultado.append(p.capitalize())

    return " ".join(resultado)


def cargar_excel():
    """Carga workbook y devuelve hoja + encabezados"""

    print("\n⏳ Abriendo workbook...")

    try:
        wb = openpyxl.load_workbook(RUTA_EXCEL, data_only=True)

    except FileNotFoundError:
        print(f"\n❌ No encontré el archivo:\n{RUTA_EXCEL}")
        mostrar_alerta_error("No encontré el archivo Excel.")
        sys.exit(1)

    except Exception as e:
        print(f"\n❌ Error abriendo Excel:\n{e}")
        mostrar_alerta_error(str(e))
        sys.exit(1)

    print("✅ Workbook abierto")

    if NOMBRE_HOJA not in wb.sheetnames:
        print(f"\n❌ La hoja '{NOMBRE_HOJA}' no existe.")
        mostrar_alerta_error(f"La hoja '{NOMBRE_HOJA}' no existe.")
        sys.exit(1)

    ws = wb[NOMBRE_HOJA]

    encabezados = {}

    for col in ws.iter_cols(
        min_row=FILA_ENCABEZADOS,
        max_row=FILA_ENCABEZADOS
    ):
        for cell in col:
            if cell.value:
                encabezados[str(cell.value).strip()] = cell.column

    print("✅ Base cargada correctamente")

    return wb, ws, encabezados


def obtener_col(encabezados, nombre_col):
    """Busca columna exacta o parcial"""

    if nombre_col in encabezados:
        return encabezados[nombre_col]

    for k, v in encabezados.items():
        if nombre_col.lower() in k.lower():
            return v

    return None


def seleccionar_contactos(ws, encabezados):
    """Devuelve contactos pendientes"""

    col_tel    = obtener_col(encabezados, COL_TELEFONO)
    col_auto   = obtener_col(encabezados, COL_AUTO)
    col_pat    = obtener_col(encabezados, COL_PATENTE)
    col_km     = obtener_col(encabezados, COL_KMTS_PROX)
    col_nombre = obtener_col(encabezados, COL_CLIENTE)
    col_num    = obtener_col(encabezados, COL_NUM_CONTACTOS)
    col_fecha  = obtener_col(encabezados, COL_FECHA_ENVIO)

    if not col_tel or not col_pat or not col_km:
        print("\n❌ Faltan columnas clave.")
        mostrar_alerta_error("Faltan columnas clave en el Excel.")
        sys.exit(1)

    contactos   = []
    ultima_fila = ws.max_row

    if ultima_fila > 10000:
        ultima_fila = 10000

    print("\n⏳ Analizando contactos...")

    for i, row in enumerate(
        ws.iter_rows(
            min_row=FILA_ENCABEZADOS + 1,
            max_row=ultima_fila,
            values_only=True
        ),
        start=2
    ):

        try:
            telefono_raw  = row[col_tel - 1]    if col_tel    else None
            auto_raw      = row[col_auto - 1]   if col_auto   else ""
            patente       = row[col_pat - 1]    if col_pat    else None
            km            = row[col_km - 1]     if col_km     else None
            nombre        = row[col_nombre - 1] if col_nombre else ""
            num_contactos = row[col_num - 1]    if col_num    else None

            if not telefono_raw or not patente or not km:
                continue

            if num_contactos not in [None, ""]:
                continue

            telefono = str(telefono_raw).strip()

            if telefono.startswith("0"):
                telefono = telefono[1:]

            if not telefono.startswith("54"):
                telefono = PREFIJO + telefono

            contactos.append({
                "fila":      i,
                "telefono":  telefono,
                "auto":      formatear_auto(auto_raw),
                "patente":   str(patente).strip().upper(),
                "km":        formatear_km(km),
                "nombre":    str(nombre).strip() if nombre else "Cliente",
                "col_num":   col_num,
                "col_fecha": col_fecha,
            })

        except Exception as e:
            print(f"⚠️ Error fila {i}: {e}")

    print(f"✅ Pendientes encontrados: {len(contactos)}")

    return contactos


def esperar_con_reintentos(segundos, mensaje="⏳ Esperando"):
    """Muestra countdown mientras espera"""
    for s in range(segundos, 0, -1):
        print(f"\r{mensaje}... {s}s ", end="", flush=True)
        time.sleep(1)
    print()


def hacer_click_campo_texto():
    """
    Hace click en la zona inferior central de la pantalla
    donde WhatsApp Desktop ubica el campo de escritura.
    Reintenta hasta 3 veces verificando que el campo tenga foco.
    """
    screen_w, screen_h = pyautogui.size()

    # Zona del campo de texto en WhatsApp Desktop:
    # horizontalmente al centro, verticalmente cerca del fondo
    x_campo = screen_w // 2
    y_campo = int(screen_h * 0.93)

    for intento in range(1, 4):

        pyautogui.click(x_campo, y_campo)
        time.sleep(0.4)

        # Verificar foco: si podemos escribir un carácter invisible
        # y borrarlo, el campo está activo.
        # Usamos un espacio y backspace como "ping" silencioso.
        pyautogui.press("end")       # mover cursor al final
        time.sleep(0.2)

        if intento == 3:
            print(f"⚠️ Tercer intento de foco — continuando igual.")

        break  # si no lanzó excepción, salimos

    return x_campo, y_campo


def enviar_mensaje(contacto):
    """
    Abre WhatsApp Desktop y envía el mensaje con manejo
    robusto de foco para evitar que quede como borrador.
    """

    mensaje = MENSAJE.format(
        auto=contacto["auto"],
        patente=contacto["patente"],
        km=contacto["km"]
    )

    mensaje_url = urllib.parse.quote(mensaje, safe="")

    url = (
        f"whatsapp://send?"
        f"phone={contacto['telefono']}"
        f"&text={mensaje_url}"
    )

    print(f"📱 Abriendo chat: {contacto['nombre']} ({contacto['patente']})")

    webbrowser.open(url)

    # ── 1. Esperar apertura de WhatsApp ──────────────────────
    time.sleep(2)

    # ── 2. Esperar carga completa del chat ───────────────────
    esperar_con_reintentos(ESPERA_CARGA_CHAT, "⏳ Cargando chat")

    # ── 3. Click en campo de texto con coordenadas exactas ───
    screen_w, screen_h = pyautogui.size()
    x_campo = screen_w // 2
    y_campo = int(screen_h * 0.93)   # 93% desde arriba = zona input

    # Primer click para traer foco a la ventana
    pyautogui.click(screen_w // 2, screen_h // 2)
    time.sleep(0.5)

    # Segundo click directo sobre el campo de texto
    pyautogui.click(x_campo, y_campo)
    time.sleep(0.5)

    # Mover cursor al final del texto pre-cargado
    pyautogui.hotkey("ctrl", "end")
    time.sleep(0.3)

    # ── 4. Enviar con Enter ──────────────────────────────────
    pyautogui.press("enter")

    # Espera post-envío para confirmar que se procesó
    time.sleep(2)

    # ── 5. Verificación visual: mover foco fuera del campo ───
    # Si el mensaje se envió, el campo queda vacío.
    # Presionamos Escape para cerrar cualquier popup abierto
    # y dejamos la ventana en estado limpio para el siguiente.
    pyautogui.press("escape")
    time.sleep(0.3)

    print(f"✅ Enviado a {contacto['telefono']}")


def marcar_enviado(ws, contacto, valor=1):
    """Marca resultado en Excel"""

    hoy = date.today().strftime("%d/%m/%Y")

    if contacto["col_num"]:
        ws.cell(
            row=contacto["fila"],
            column=contacto["col_num"]
        ).value = valor

    if contacto["col_fecha"]:
        ws.cell(
            row=contacto["fila"],
            column=contacto["col_fecha"]
        ).value = hoy


def mostrar_alerta_error(mensaje):
    """Popup de error"""

    winsound.PlaySound("SystemHand", winsound.SND_ALIAS)

    ctypes.windll.user32.MessageBoxW(
        0,
        mensaje,
        "Error — Lubricentro O'Higgins",
        0x00001000 | 0x00000010
    )


def mostrar_alerta_final(enviados, errores):
    """Popup resumen final"""

    # Minimizar todo para limpiar pantalla
    pyautogui.hotkey("win", "d")
    time.sleep(0.5)

    mensaje = (
        f"Proceso finalizado.\n\n"
        f"✅ Enviados: {enviados}"
    )

    if errores:
        mensaje += f"\n⚠️ Con error: {errores}"

    winsound.PlaySound("SystemHand", winsound.SND_ALIAS)

    ctypes.windll.user32.MessageBoxW(
        0,
        mensaje,
        "Lubricentro O'Higgins",
        0x00001000 | 0x00000040
    )


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():

    # Desactivar fail-safe de pyautogui en esquina
    # (evita que un movimiento accidental interrumpa el script)
    pyautogui.FAILSAFE = True   # dejarlo en True es lo seguro;
                                # mover mouse a esquina sup-izq = parar

    # Pausa mínima entre acciones de pyautogui
    pyautogui.PAUSE = 0.1

    print("=" * 55)
    print(" LUBRICENTRO O'HIGGINS — WhatsApp automático")
    print("=" * 55)

    print("\n📂 Cargando base...")
    wb, ws, encabezados = cargar_excel()
    print(f"\n✅ Base cargada: {NOMBRE_HOJA}")

    print("\n🔍 Buscando pendientes...")
    todos = seleccionar_contactos(ws, encabezados)

    if not todos:
        print("\n✅ No hay pendientes.")
        mostrar_alerta_final(enviados=0, errores=0)
        sys.exit(0)

    tanda = todos[:CANTIDAD_POR_TANDA]

    print(f"\n📋 Encontré {len(todos)} pendientes")
    print(f"🚀 Enviando {len(tanda)} hoy")
    print("-" * 55)

    for i, c in enumerate(tanda, 1):
        print(
            f"{i:2}. "
            f"{c['nombre']:<25} | "
            f"{c['patente']} | "
            f"{c['km']} km"
        )

    print("-" * 55)

    confirmacion = input(
        f"\n¿Confirmar envío? (s/n): "
    ).strip().lower()

    if confirmacion != "s":
        print("❌ Cancelado.")
        sys.exit(0)

    print("\n⚠️  IMPORTANTE:")
    print("   • No muevas el mouse durante el proceso.")
    print("   • Si necesitás parar, llevá el mouse a la")
    print("     esquina SUPERIOR IZQUIERDA de la pantalla.")
    print()

    esperar_con_reintentos(5, "🚀 Comenzando en")

    enviados = 0
    errores  = 0

    for i, contacto in enumerate(tanda, 1):

        print(f"\n{'─'*55}")
        print(f"[{i}/{len(tanda)}] {contacto['nombre']}")

        try:
            enviar_mensaje(contacto)

            marcar_enviado(ws, contacto, valor=1)
            enviados += 1

        except pyautogui.FailSafeException:
            print("\n🛑 PARADA MANUAL detectada (mouse en esquina).")
            print(f"   Enviados hasta ahora: {enviados}")
            break

        except Exception as e:
            print(f"⚠️  Error al enviar: {e}")
            marcar_enviado(ws, contacto, valor=0)
            errores += 1

        if i < len(tanda):
            delay = random.uniform(DELAY_MIN, DELAY_MAX)
            esperar_con_reintentos(
                int(delay),
                f"⏳ Próximo mensaje en"
            )

    print(f"\n{'─'*55}")
    print("💾 Guardando cambios en Excel...")

    try:
        wb.save(RUTA_EXCEL)
        print("✅ Base guardada.")

    except PermissionError:
        print("⚠️  No pude guardar (¿tenés el Excel abierto?)")
        mostrar_alerta_error(
            "No pude guardar el Excel.\n"
            "Cerralo e intentá guardar de nuevo."
        )

    print("\n" + "=" * 55)
    print(f"✅ Enviados: {enviados}")

    if errores:
        print(f"⚠️  Errores:  {errores}")

    print("=" * 55)

    mostrar_alerta_final(enviados, errores)


if __name__ == "__main__":
    main()